import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from tkinter import Tk, messagebox, simpledialog
from tkinter.filedialog import askdirectory, asksaveasfilename, askopenfilename

# 创建一个函数，让用户打开指定文件夹，遍历该文件夹中的所有excel文件
# 文件夹由用户指定，弹出一个文件管理器，让用户自己去选择文件夹
def select_folder():
    root = Tk()
    root.withdraw()  # we don't want a full GUI, so keep the root window from appearing
    folder_path = askdirectory(title="Select Folder Containing Excel Files")
    root.destroy()
    return folder_path

# 在遍历文件夹中的excel文件时，应该去掉 类似这种的临时文件 ~$0953-DX-挥发酚.xlsx
def get_excel_files_from_folder(folder_path):
    excel_files = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xls'):
                if not file.startswith('~$'):
                    excel_files.append(os.path.join(root, file))
    return excel_files

def _normalize_sample_id(value: object) -> str:
    """
    将样品编号做规范化，便于匹配：
    - 转大写
    - 去除空白、全角空白
    - 保留"平行X"字样（作为独立样品编号）
    - 去除中文括号及常见标点
    """
    s = str(value) if value is not None else ""
    s = s.upper().strip()
    # 去掉空白
    s = re.sub(r"\s+", "", s)
    # 注意：不再去除"平行1/2/..."字样，保留作为独立样品编号
    # s = re.sub(r"平行[0-9]+", "", s)  # 已禁用
    # 去掉中英文括号及其中的空内容痕迹
    s = s.replace("（", "(").replace("）", ")").replace("【", "[").replace("】", "]")
    s = s.replace("，", ",").replace("。", ".")
    return s

def _expand_range_pattern(pattern: str):
    """
    展开区间模式的样品编号。
    
    支持格式:
    - DX2523660(1-6)01 → 英文括号，展开为 DX2523660101, DX2523660201, ..., DX2523660601
    - DX2523660（1-6）01 → 中文括号，展开为 DX2523660101, DX2523660201, ..., DX2523660601
    - ABC(1-3)XYZ → ABC1XYZ, ABC2XYZ, ABC3XYZ
    - TEST（01-05） → TEST01, TEST02, TEST03, TEST04, TEST05
    
    参数:
        pattern: 可能包含区间的样品编号模式
    
    返回:
        展开后的编号列表，如果不是区间模式则返回原编号
    """
    # 匹配模式1: 英文括号 前缀(起始-结束)后缀
    match = re.match(r'^(.*?)\((\d+)-(\d+)\)(.*)$', pattern.strip())
    
    if not match:
        # 匹配模式2: 中文括号 前缀（起始-结束）后缀
        match = re.match(r'^(.*?)（(\d+)-(\d+)）(.*)$', pattern.strip())
    
    if not match:
        # 不是区间模式，返回原字符串
        return [pattern]
    
    prefix = match.group(1)  # 前缀（区间前的部分）
    start_str = match.group(2)  # 起始数字
    end_str = match.group(3)  # 结束数字
    suffix = match.group(4)  # 后缀（区间后的部分）
    
    # 获取数字的位数（用于补零）
    width = len(start_str)
    
    try:
        start = int(start_str)
        end = int(end_str)
        
        if start > end:
            print(f"⚠ 警告: 区间起始值({start})大于结束值({end})，已自动交换")
            start, end = end, start
        
        # 生成区间内的所有编号
        result = []
        for i in range(start, end + 1):
            # 格式化数字，保持位数（补零）
            num_str = str(i).zfill(width)
            # 组合: 前缀 + 数字 + 后缀
            result.append(f"{prefix}{num_str}{suffix}")
        
        return result
    except ValueError:
        print(f"⚠ 警告: 无法解析区间 '{pattern}'，作为普通编号处理")
        return [pattern]

def _parse_target_ids(text: str):
    """
    将文本字符串解析为编号列表。
    支持逗号/空格/分号/中文逗号/换行等分隔符，自动去重并转大写。
    支持区间语法(英文或中文括号): 
    - DX2523660(1-6)01 或 DX2523660（1-6）01 → 展开为 DX2523660101, 201, 301, 401, 501, 601
    - ABC(1-5) 或 ABC（1-5） → 展开为 ABC1, ABC2, ABC3, ABC4, ABC5
    - TEST(01-05)XYZ 或 TEST（01-05）XYZ → 展开为 TEST01XYZ, TEST02XYZ, ..., TEST05XYZ
    
    参数:
        text: 多个编号的文本（以分隔符分开）
    
    返回:
        去重后的编号列表
    """
    if not text:
        return []
    parts = re.split(r"[，,;；\s]+", text.strip())
    ids = []
    seen = set()
    for p in parts:
        p = p.strip()
        if not p:
            continue
        
        # 展开可能的区间模式
        expanded = _expand_range_pattern(p)
        
        for item in expanded:
            item_upper = item.upper()
            if item_upper not in seen:
                seen.add(item_upper)
                ids.append(item_upper)
    return ids

def read_ids_from_txt(file_path: str):
    """
    从TXT文件读取编号列表，每行一个编号。
    自动过滤空行和注释行，去重并转大写。
    支持区间语法(英文或中文括号): 
    - DX2523660(1-6)01 或 DX2523660（1-6）01 → 展开为 DX2523660101, 201, 301, 401, 501, 601
    - ABC(1-5) 或 ABC（1-5） → 展开为 ABC1, ABC2, ABC3, ABC4, ABC5
    
    参数:
        file_path: TXT文件的完整路径
    
    返回:
        去重后的编号列表，若读取失败返回 None
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        ids = []
        seen = set()
        for line in lines:
            line = line.strip()
            # 跳过空行和注释行（以#开头）
            if not line or line.startswith('#'):
                continue
            
            # 展开可能的区间模式
            expanded = _expand_range_pattern(line)
            
            for item in expanded:
                item_upper = item.upper()
                if item_upper not in seen:
                    seen.add(item_upper)
                    ids.append(item_upper)
        
        return ids if ids else None
    except Exception as e:
        messagebox.showerror("读取失败", f"无法读取TXT文件：\n{e}")
        return None

def extract_metadata_from_excel(file_path, sheet_name=None):
    """
    从Excel文件中提取元数据信息
    通过查找这些标签名称，然后获取其右侧相邻单元格的值
    
    参数:
        file_path: Excel文件的完整路径
        sheet_name: 可选，指定要读取的sheet名称。若不提供，则读取第一个sheet。
    
    返回:
        字典，包含 {
            'analyzer': '分析人/分析人员',
            'instrument_number': '仪器编号',
            'analysis_method': '分析方法',
            'detection_limit': '检出限',
            'analysis_date': '分析日期'
        }
    """
    metadata = {
        'analyzer': '',
        'instrument_name': '',  # 使用仪器
        'instrument_number': '',  # 仪器编号
        'analysis_method': '',
        'detection_limit': '',
        'analysis_date': ''
    }
    
    try:
        if file_path.endswith('.xlsx'):
            wb = load_workbook(file_path, data_only=True)
            if not wb.sheetnames:
                return metadata
            
            # 选择指定的sheet或默认第一个sheet
            if sheet_name is not None:
                ws = wb[sheet_name]
            else:
                ws = wb[wb.sheetnames[0]]
            
            # 标签和对应字段的映射（包含多个可能的标签名称）
            label_mapping = {
                '分析人': 'analyzer',
                '分析人员': 'analyzer',
                '使用仪器': 'instrument_name',
                '仪器编号': 'instrument_number',
                '仪器型号': 'instrument_number',
                '分析方法': 'analysis_method',
                '检出限': 'detection_limit',
                '分析日期': 'analysis_date'
            }
            
            # 遍历所有单元格查找标签
            for row in ws.iter_rows(min_row=1, max_row=100):
                for cell in row:
                    cell_value = str(cell.value).strip() if cell.value else ""
                    if cell_value in label_mapping:
                        # 找到标签，获取右侧相邻单元格的值
                        right_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        field_name = label_mapping[cell_value]
                        # 只在该字段还没有值时才填充（优先使用先找到的值）
                        if not metadata[field_name]:
                            metadata[field_name] = str(right_cell.value).strip() if right_cell.value else ""
        else:
            # .xls 文件处理
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            label_mapping = {
                '分析人': 'analyzer',
                '分析人员': 'analyzer',
                '使用仪器': 'instrument_name',
                '仪器编号': 'instrument_number',
                '仪器型号': 'instrument_number',
                '分析方法': 'analysis_method',
                '检出限': 'detection_limit',
                '分析日期': 'analysis_date'
            }
            
            for idx, row in df.iterrows():
                for col_idx, cell_value in enumerate(row):
                    cell_str = str(cell_value).strip() if cell_value and pd.notna(cell_value) else ""
                    if cell_str in label_mapping:
                        # 获取右侧相邻单元格
                        if col_idx + 1 < len(row):
                            right_value = row.iloc[col_idx + 1]
                            field_name = label_mapping[cell_str]
                            # 只在该字段还没有值时才填充
                            if not metadata[field_name]:
                                metadata[field_name] = str(right_value).strip() if right_value and pd.notna(right_value) else ""
    except Exception as e:
        print(f"⚠ 提取元数据失败: {e}")
    
    return metadata

def ask_sample_ids_source():
    """
    弹出对话框让用户选择样品编号的输入方式：
    1. 直接输入 - 在对话框中手动输入多个编号
    2. 从TXT文件导入 - 选择一个TXT文件，每行一个编号
    
    返回:
        编号列表，若用户取消则返回 None
    """
    root = Tk()
    root.withdraw()
    
    # 先让用户选择输入方式
    choice = messagebox.askyesnocancel(
        title="选择样品编号输入方式",
        message="请选择如何输入样品编号：\n"
                "\n"
                "\"是\" - 直接输入：在对话框中输入多个编号（逗号/空格/分号分隔）\n"
                "\"否\" - 从文件导入：选择一个TXT文件，每行一个编号\n"
                "\"取消\" - 退出程序",
    )
    
    if choice is None:  # 取消
        root.destroy()
        return None
    elif choice:  # 是 - 直接输入
        ids_text = simpledialog.askstring(
            title="输入样品编号",
            prompt=(
                "请输入要提取的样品编号：\n"
                "- 多个编号可用逗号/空格/分号/中文逗号/换行分隔\n"
                "- 区间语法: DX252366(01-06)01 或 DX252366（01-06）01\n"
                "- 例如：DX2509530201, DX252366(01-06)01"
            ),
            initialvalue="DX2509530201, DX2509530301",
            parent=root,
        )
        root.destroy()
        if ids_text:
            return _parse_target_ids(ids_text)
        else:
            return None
    else:  # 否 - 从文件导入
        root.destroy()
        file_path = askopenfilename(
            title="选择包含样品编号的TXT文件",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )
        if file_path:
            return read_ids_from_txt(file_path)
        else:
            return None

def get_sheets_to_process(file_path):
    """
    获取需要处理的sheet页列表及对应的项目名提取策略
    
    参数:
        file_path: Excel文件的完整路径
    
    返回:
        list of dict: [
            {
                'sheet_name': sheet名称,
                'project_name_source': 'sheet' 或 'file' (项目名来源)
            },
            ...
        ]
    """
    try:
        # 读取所有sheet名称
        if file_path.endswith('.xlsx'):
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        else:  # .xls
            xls = pd.ExcelFile(file_path)
            sheet_names = xls.sheet_names
            xls.close()
        
        # 判断sheet名称模式
        # 检查是否所有sheet都符合 "sheet[1-9]" 模式(不区分大小写)
        sheet_pattern = re.compile(r'^sheet[1-9]$', re.IGNORECASE)
        all_match_pattern = all(sheet_pattern.match(name) for name in sheet_names)
        
        result = []
        if all_match_pattern:
            # 情况2: 所有sheet都是sheet[1-9]格式，只处理sheet1，项目名从文件名获取
            for name in sheet_names:
                if name.lower() == 'sheet1':
                    result.append({
                        'sheet_name': name,
                        'project_name_source': 'file'
                    })
                    break
        else:
            # 情况1: 存在非sheet[1-9]格式的sheet，处理所有非sheet[1-9]的sheet，项目名从sheet名获取
            for name in sheet_names:
                if not sheet_pattern.match(name):
                    result.append({
                        'sheet_name': name,
                        'project_name_source': 'sheet'
                    })
        
        return result
    except Exception as e:
        print(f"⚠ 读取sheet列表失败: {e}")
        # 出错时返回默认第一个sheet，项目名从文件名获取
        return [{'sheet_name': 0, 'project_name_source': 'file'}]

def extract_sample_and_concentration(file_path, skip_empty_rows=False, targets=None, sheet_name=None):
    """
    提取Excel文件中的样品编号和浓度列数据
    
    参数:
        file_path: Excel文件的完整路径
        skip_empty_rows: 是否跳过样品编号和浓度都为空的行 (默认False)
        targets: 可选，目标样品编号列表（例如 ["DX2509530201", "DX2509530301"]）。
                 若提供，则只返回匹配这些样品编号的行。
        sheet_name: 可选，指定要读取的sheet名称或索引。若不提供，则读取第一个sheet。
    
    返回:
        二维数组，每行包含[样品编号, 浓度]
    """
    print(f"\n{'='*80}")
    print(f"处理文件: {file_path}")
    if sheet_name is not None:
        print(f"处理Sheet: {sheet_name}")
    print(f"{'='*80}")
    
    # 样品编号和浓度列名的可能变体
    sample_col_names = ['样品编号', '编号']
    concentration_col_names = ['样品浓度', '计算结果浓度']  # 仅提取这两个列名
    
    df = None
    header_row = None
    sample_col = None
    concentration_col = None
    
    # 尝试不同的header行 (0-99，即第1-100行) 来找到包含目标列的行
    for header in range(100):
        try:
            temp_df = pd.read_excel(file_path, sheet_name=sheet_name, header=header)
        
            # 查找样品编号列
            temp_sample_col = None
            for col_name in sample_col_names:
                if col_name in temp_df.columns:
                    temp_sample_col = col_name
                    break
            
            # 查找浓度列（精确匹配"样品浓度"或"计算结果浓度"，或模糊匹配包含这些关键词的列名）
            temp_concentration_col = None
            for col_name in temp_df.columns:
                # 先尝试精确匹配
                if col_name in concentration_col_names:
                    temp_concentration_col = col_name
                    break
                # 再尝试模糊匹配（包含"样品浓度"或"计算结果浓度"）
                for keyword in concentration_col_names:
                    if keyword in col_name:
                        temp_concentration_col = col_name
                        break
                if temp_concentration_col:
                    break
            
            # 如果找到了两个目标列，就使用这个header
            if temp_sample_col and temp_concentration_col:
                df = temp_df
                header_row = header
                sample_col = temp_sample_col
                concentration_col = temp_concentration_col
                print(f"✓ 使用第 {header+1} 行作为列名 (header={header})")
                break
        except Exception as e:
            error_msg = str(e)
            # 如果是 xlrd 相关错误，说明文件是 .xls 格式但缺少 xlrd 库
            if 'xlrd' in error_msg.lower():
                print(f"❌ 文件格式错误: 该文件是 .xls 格式，需要安装 xlrd 库来支持")
                print(f"   解决方案: pip install xlrd 或 uv add xlrd")
                return []
            # 其他异常继续尝试
            print(f"尝试 header={header} 失败: {e}")
            continue
    # 如果没有找到合适的列
    if df is None or sample_col is None or concentration_col is None:
        print(f"\n⚠ 警告: 未找到样品编号或浓度列")
        if df is not None:
            print(f"可用的列名: {list(df.columns)}")
        return []
    
    print(f"✓ 找到样品编号列: '{sample_col}'")
    print(f"✓ 找到浓度列: '{concentration_col}'")
    
    # 对于 .xlsx 文件，读取单元格的显示值（应用格式化）
    result_array = []
    if file_path.endswith('.xlsx'):
        try:
            # 同时加载两个工作簿版本
            wb_values = load_workbook(file_path, data_only=True)  # 获取缓存的计算结果
            wb_format = load_workbook(file_path, data_only=False)  # 获取格式信息
            
            # 选择指定的sheet或默认第一个sheet
            if sheet_name is not None:
                ws_values = wb_values[sheet_name]
                ws_format = wb_format[sheet_name]
            else:
                ws_values = wb_values[wb_values.sheetnames[0]]
                ws_format = wb_format[wb_format.sheetnames[0]]
            
            # 找到样品编号和浓度列在工作表中的列索引
            sample_col_idx = None
            concentration_col_idx = None
            
            # 遍历header行找到列索引（从1开始）
            for col_idx, cell in enumerate(ws_format.iter_cols(min_row=header_row+1, max_row=header_row+1, values_only=False), 1):
                if cell[0].value == sample_col:
                    sample_col_idx = col_idx
                elif cell[0].value == concentration_col:
                    concentration_col_idx = col_idx
            
            # 从工作表中逐行读取值并应用格式化
            if sample_col_idx and concentration_col_idx:
                for row_idx in range(header_row + 2, ws_values.max_row + 1):
                    sample_cell = ws_values.cell(row=row_idx, column=sample_col_idx)
                    concentration_cell_values = ws_values.cell(row=row_idx, column=concentration_col_idx)
                    concentration_cell_format = ws_format.cell(row=row_idx, column=concentration_col_idx)
                    
                    sample = sample_cell.value
                    concentration = concentration_cell_values.value
                    
                    # 对浓度值应用单元格格式化（获取Excel显示值）
                    if concentration is not None and isinstance(concentration, (int, float)):
                        num_format = concentration_cell_format.number_format
                        
                        # 调试：打印格式信息（只打印前几行）
                        if row_idx <= header_row + 5:
                            print(f"调试 - 行{row_idx}: 原值={concentration}, 格式='{num_format}'")
                        
                        # 根据单元格格式进行智能格式化，模拟Excel显示逻辑
                        if num_format == '0' or num_format == '#,##0':  # 整数格式
                            concentration = int(round(concentration))
                        elif '.' in num_format:  # 包含小数点的格式
                            try:
                                # 提取小数点后的部分，解析小数位数
                                after_dot = num_format.split('.')[-1]
                                # 计算小数位数（统计0和#的个数）
                                decimals = 0
                                for char in after_dot:
                                    if char in ('0', '#'):
                                        decimals += 1
                                    elif char not in (' ', '_', ';', '-', ','):
                                        break  # 遇到其他字符停止计数
                                
                                if decimals > 0:
                                    # 应用四舍五入到指定小数位数
                                    concentration = round(concentration, decimals)
                            except:
                                pass  # 解析失败，保持原值
                        elif num_format == 'General' or not num_format:
                            # General格式或无格式，保持原值
                            pass
                        
                        if row_idx <= header_row + 5:
                            print(f"         转换后={concentration}")
                    
                    # 如果需要跳过空行
                    if skip_empty_rows:
                        if sample is None and concentration is None:
                            continue
                    
                    result_array.append([sample, concentration])
        except Exception as e:
            print(f"⚠ 使用 openpyxl 读取失败，回退到 pandas: {e}")
            import traceback
            traceback.print_exc()
            # 回退到 pandas 读取
            sample_data = df[sample_col].tolist()
            concentration_data = df[concentration_col].tolist()
            for sample, concentration in zip(sample_data, concentration_data):
                if skip_empty_rows:
                    if pd.isna(sample) and pd.isna(concentration):
                        continue
                result_array.append([sample, concentration])
    else:
        # .xls 文件使用 xlrd 直接读取并应用格式化（获取Excel显示值）
        print("  使用 xlrd 读取 .xls 文件并应用单元格格式...")
        try:
            import xlrd
            # 使用 formatting_info=True 来获取格式信息
            wb_xls = xlrd.open_workbook(file_path, formatting_info=True)
            ws_xls = wb_xls.sheet_by_name(sheet_name) if sheet_name else wb_xls.sheet_by_index(0)
            
            # 找到样品编号和浓度列的索引
            sample_col_idx = None
            concentration_col_idx = None
            for col_idx in range(ws_xls.ncols):
                cell_value = ws_xls.cell_value(header_row, col_idx)
                if cell_value == sample_col:
                    sample_col_idx = col_idx
                elif cell_value == concentration_col:
                    concentration_col_idx = col_idx
            
            if sample_col_idx is None or concentration_col_idx is None:
                print("  ⚠ 在 .xls 文件中未找到目标列，回退到 pandas")
                # 回退到pandas
                sample_data = df[sample_col].tolist()
                concentration_data = df[concentration_col].tolist()
                for sample, concentration in zip(sample_data, concentration_data):
                    if skip_empty_rows:
                        if pd.isna(sample) and pd.isna(concentration):
                            continue
                    result_array.append([sample, concentration])
            else:
                # 从数据行开始读取（header_row + 1）
                for row_idx in range(header_row + 1, ws_xls.nrows):
                    sample = ws_xls.cell_value(row_idx, sample_col_idx)
                    concentration = ws_xls.cell_value(row_idx, concentration_col_idx)
                    
                    # 对浓度值应用格式化（获取Excel显示值）
                    if isinstance(concentration, (int, float)):
                        try:
                            # 获取单元格格式索引
                            xf_index = ws_xls.cell_xf_index(row_idx, concentration_col_idx)
                            xf = wb_xls.format_map.get(wb_xls.xf_list[xf_index].format_key)
                            
                            if xf:
                                format_str = xf.format_str
                                
                                # 调试：打印前几行的格式信息
                                if row_idx <= header_row + 5:
                                    print(f"  调试 - 行{row_idx}: 原值={concentration}, 格式='{format_str}'", end="")
                                
                                # 根据格式字符串进行格式化
                                # 小数格式（优先判断，因为包含小数点）
                                if '.' in format_str:
                                    try:
                                        # 提取小数点后的部分
                                        after_dot = format_str.split('.')[-1]
                                        # 计算小数位数（统计0和#的个数）
                                        decimals = 0
                                        for char in after_dot:
                                            if char in ('0', '#'):
                                                decimals += 1
                                            elif char not in (' ', '_', ';', '-', ',', ')'):
                                                break
                                        
                                        if decimals > 0:
                                            # 格式化为字符串以保持小数位数（包括末尾的0）
                                            # 例如：7.9 + 格式'0.00' → "7.90"
                                            format_template = f"{{:.{decimals}f}}"
                                            concentration = format_template.format(concentration)
                                            # 保持为字符串，这样才能保留末尾的0
                                        else:
                                            # 小数点后没有0或#，当作整数处理
                                            concentration = int(round(concentration))
                                    except:
                                        pass
                                # 整数格式（包含0或#但不包含小数点）
                                elif '0' in format_str or '#' in format_str:
                                    concentration = int(round(concentration))
                                # General格式，保持原值
                                elif format_str == 'General':
                                    pass
                                
                                if row_idx <= header_row + 5:
                                    print(f" → 转换后={concentration}")
                        except:
                            pass
                    
                    # 如果需要跳过空行
                    if skip_empty_rows:
                        if (sample is None or sample == '') and (concentration is None or concentration == ''):
                            continue
                    
                    result_array.append([sample, concentration])
        
        except Exception as e:
            print(f"  ⚠ 使用 xlrd 读取格式失败: {e}，回退到 pandas")
            # 回退到pandas读取
            sample_data = df[sample_col].tolist()
            concentration_data = df[concentration_col].tolist()
            for sample, concentration in zip(sample_data, concentration_data):
                if skip_empty_rows:
                    if pd.isna(sample) and pd.isna(concentration):
                        continue
                result_array.append([sample, concentration])
    
    # 注意：平行样品现在作为独立样品处理
    # - 不再填充平行样品的空值
    # - 不再去除"平行X"字样
    # - 浓度为空时保持空值(NaN)
    
    # 筛选符合条件的行，形成新数组
    filtered_array = result_array.copy()
    #for i, row in enumerate(result_array):
    #    # 只保留包含250953且不包含KB、PS的行
    #    if '250953' in str(row[0]) and 'KB' not in str(row[0]) and 'PS' not in str(row[0]):
    #        filtered_array.append(row)

    # 如果指定了目标样品编号，则进一步按目标过滤（按规范化后精确匹配）
    if targets:
        normalized_targets = {_normalize_sample_id(t) for t in targets}
        filtered_array = [
            [row[0], row[1]]
            for row in filtered_array
            if _normalize_sample_id(row[0]) in normalized_targets
        ]
    
    # 打印结果
    print(f"\n提取的二维数组 (共 {len(result_array)} 行):")
    if skip_empty_rows:
        print("(已过滤掉空行)")
    print(f"列顺序: [样品编号, 浓度]")
    
    print(f"\n符合条件的数据 (共 {len(filtered_array)} 行):")
    print("(包含250953 且 不包含KB、PS)")
    for i, row in enumerate(filtered_array):
        print(f"* 行 {i}: {row}")
    
    # 返回筛选后的数组
    return filtered_array

if __name__ == "__main__":
    folder_path = select_folder()
    
    # 让用户选择样品编号输入方式：直接输入或从TXT文件导入
    target_sample_ids = ask_sample_ids_source()
    if not target_sample_ids:
        messagebox.showwarning("未输入编号", "未输入任何样品编号，程序将退出。")
        raise SystemExit(0)

    excel_files = get_excel_files_from_folder(folder_path)
    print("Found Excel files:")
    print(f"待处理的样品编号（共 {len(target_sample_ids)} 个）: {', '.join(target_sample_ids)}")
    
    # 创建一个总的数组来存储所有项目的数据
    all_data = []
    
    for excel_file in excel_files:
        print(excel_file)
        
        # 获取需要处理的sheet列表
        sheets_info = get_sheets_to_process(excel_file)
        
        # 获取文件名并提取汉字部分（用于项目名从文件名获取的情况）
        file_name = os.path.basename(excel_file)
        chinese_chars = re.findall(r'[\u4e00-\u9fff]+', file_name)
        file_chinese_name = ''.join(chinese_chars) if chinese_chars else file_name
        
        # 处理每个sheet
        for sheet_info in sheets_info:
            sheet_name = sheet_info['sheet_name']
            project_name_source = sheet_info['project_name_source']
            
            # 为每个sheet提取元数据（从当前sheet页中提取）
            metadata = extract_metadata_from_excel(excel_file, sheet_name=sheet_name)
            
            # 提取数据
            result_array = extract_sample_and_concentration(
                excel_file,
                skip_empty_rows=True,
                targets=target_sample_ids,
                sheet_name=sheet_name,
            )
            
            # 根据来源确定项目名
            if project_name_source == 'sheet':
                # 从sheet名称提取汉字
                sheet_chinese_chars = re.findall(r'[\u4e00-\u9fff]+', str(sheet_name))
                project_name = ''.join(sheet_chinese_chars) if sheet_chinese_chars else str(sheet_name)
            else:  # 'file'
                # 从文件名提取汉字
                project_name = file_chinese_name
            
            # 基于result_array，添加元数据列
            for row in result_array:
                # 仪器编号列：优先使用 instrument_number，如果为空则使用 instrument_name
                # 这样可以处理两种情况：
                # 1. 有"仪器编号"标签 → 使用仪器编号的值（如 /）
                # 2. 只有"使用仪器"标签 → 使用使用仪器的值（如 酸式滴定管）
                instrument_info = metadata.get('instrument_number', '') or metadata.get('instrument_name', '')
                
                # 列顺序：分析人、仪器编号、分析方法、检出限、分析日期、项目名、样品编号、样品浓度
                row.insert(0, metadata.get('analysis_date', ''))
                row.insert(0, metadata.get('detection_limit', ''))
                row.insert(0, metadata.get('analysis_method', ''))
                row.insert(0, instrument_info)
                row.insert(0, metadata.get('analyzer', ''))
                row.insert(5, project_name)  # 项目名在第6列（索引5）
                all_data.append(row)
    
    # 将所有数据写入到一个Excel文件中
    if all_data:
        # 创建DataFrame，包含新增的元数据列
        df = pd.DataFrame(all_data, columns=[
            "分析人", 
            "仪器编号",
            "分析方法",
            "检出限",
            "分析日期", 
            "项目名", 
            "样品编号", 
            "样品浓度"
        ])
        
        # 弹出文件保存对话框（改为：先选路径，再输入不带后缀的文件名）
        root = Tk()
        root.withdraw()  # 隐藏主窗口

        # 默认文件名（包含关键词以便区分）
        default_filename = "汇总_0953_指定样品提取.xlsx"
        default_basename = os.path.splitext(default_filename)[0]

        # 1) 选择保存文件夹
        messagebox.showinfo(
            "保存提取结果",
            (
                f"即将保存提取结果：\n\n"
                f"数据行数: {len(all_data)} 行\n"
                f"来源文件: {len(excel_files)} 个\n\n"
                f"请先选择保存文件夹，然后输入文件名（不含后缀）。"
            ),
        )
        output_dir = askdirectory(title="选择保存文件夹", initialdir=folder_path)
        if not output_dir:
            print("\n⚠ 用户取消了保存操作（未选择保存文件夹）")
            messagebox.showwarning("取消保存", "未选择保存文件夹，未保存提取结果")
            raise SystemExit(0)

        # 2) 输入不带后缀的文件名
        filename_input = simpledialog.askstring(
            title="输入文件名",
            prompt="请输入要保存的文件名（不含后缀 .xlsx）：",
            initialvalue=default_basename,
            parent=root,
        )

        base_name = (filename_input or "").strip()
        if not base_name:
            messagebox.showinfo("使用默认文件名", f"未输入文件名，将使用默认文件名：{default_filename}")
            base_name = default_basename

        # 3) 组装最终路径并覆盖确认
        output_file = os.path.join(output_dir, f"{base_name}.xlsx")
        if os.path.exists(output_file):
            from tkinter import messagebox as _mb
            if not _mb.askyesno("文件已存在", f"{output_file}\n已存在，是否覆盖？"):
                print("\n⚠ 用户取消了保存操作（拒绝覆盖已存在文件）")
                messagebox.showwarning("取消保存", "未保存提取结果")
                raise SystemExit(0)

        # 4) 保存
        try:
            df.to_excel(output_file, index=False)
            print(f"\n{'='*80}")
            print(f"✅ 所有提取结果已汇总保存到: {output_file}")
            print(f"   共提取 {len(all_data)} 行数据，来自 {len(excel_files)} 个文件")
            print(f"{'='*80}")
            
            # 显示成功提示
            messagebox.showinfo("保存成功", f"提取结果已成功保存到:\n{output_file}")
        except Exception as e:
            print(f"\n❌ 保存失败: {e}")
            messagebox.showerror("保存失败", f"保存失败:\n{e}")
    else:
        print("\n⚠ 没有提取到任何数据")
import os
from openpyxl import load_workbook, Workbook
import xlrd
from xlrd import formatting

def format_cell_value(cell_obj, datemode):
    """
    根据单元格格式返回格式化后的显示值
    """
    cell_value = cell_obj.value
    cell_type = cell_obj.ctype
    
    # 0=空, 1=文本, 2=数字, 3=日期, 4=布尔, 5=错误, 6=空白
    if cell_type == 0 or cell_type == 6:  # 空单元格
        return None
    elif cell_type == 1:  # 文本
        return cell_value
    elif cell_type == 2:  # 数字
        return cell_value  # xlrd已经返回了Python数值类型
    elif cell_type == 3:  # 日期
        try:
            date_tuple = xlrd.xldate_as_tuple(cell_value, datemode)
            return f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
        except:
            return cell_value
    elif cell_type == 4:  # 布尔
        return bool(cell_value)
    elif cell_type == 5:  # 错误
        return f"#ERROR{cell_value}"
    else:
        return cell_value


def read_xls_with_formatting(filepath, sheet_index=0):
    """
    读取.xls文件并应用格式化，返回显示值（肉眼看到的值）
    """
    # 使用 formatting_info=True 来获取格式信息
    wb = xlrd.open_workbook(filepath, formatting_info=True)
    ws = wb.sheet_by_index(sheet_index)
    
    all_data = []
    
    print(f"\n{'='*80}")
    print(f"文件: {os.path.basename(filepath)}")
    print(f"Sheet: {ws.name}")
    print(f"总行数: {ws.nrows}, 总列数: {ws.ncols}")
    print(f"{'='*80}\n")
    
    for row_idx in range(ws.nrows):
        row_data = []
        for col_idx in range(ws.ncols):
            cell = ws.cell(row_idx, col_idx)
            cell_value = cell.value
            
            # 对于数字类型，尝试应用格式
            if cell.ctype == 2:  # 数字类型
                try:
                    # 获取单元格格式索引
                    xf_index = ws.cell_xf_index(row_idx, col_idx)
                    xf = wb.format_map.get(wb.xf_list[xf_index].format_key)
                    
                    if xf:
                        format_str = xf.format_str
                        
                        # 打印前几行的格式信息（调试用）
                        if row_idx < 20 and col_idx < 10:
                            print(f"单元格({row_idx},{col_idx}): 原值={cell_value}, 格式='{format_str}'", end="")
                        
                        # 根据格式字符串进行格式化
                        if format_str:
                            # 小数格式（优先判断，因为包含小数点）
                            if '.' in format_str:
                                # 计算小数位数
                                try:
                                    after_dot = format_str.split('.')[-1]
                                    decimals = 0
                                    for char in after_dot:
                                        if char in ('0', '#'):
                                            decimals += 1
                                        elif char not in (' ', '_', ';', '-', ',', ')'):
                                            break
                                    
                                    if decimals > 0:
                                        cell_value = round(cell_value, decimals)
                                    else:
                                        # 小数点后没有0或#，当作整数处理
                                        cell_value = int(round(cell_value))
                                except:
                                    pass
                            # 整数格式（包含0或#但不包含小数点）
                            elif '0' in format_str or '#' in format_str:
                                cell_value = int(round(cell_value))
                            # General格式，保持原值
                            elif format_str == 'General':
                                pass
                        
                        if row_idx < 20 and col_idx < 10:
                            print(f" → 转换后={cell_value}")
                    
                except Exception as e:
                    if row_idx < 5:
                        print(f"  格式化失败: {e}")
                    pass
            
            row_data.append(cell_value)
        
        all_data.append(row_data)
    
    return all_data


def copy_visible_excel(src_path, dest_path="visible_only.xlsx"):
    ext = os.path.splitext(src_path)[1].lower()
    new_wb = Workbook()
    new_ws = new_wb.active

    if ext in [".xlsx", ".xlsm"]:
        print("读取 XLSX 格式文件 ...")
        wb = load_workbook(src_path, data_only=True)
        ws = wb.active

        dest_row = 1
        for row in ws.iter_rows():
            # 检查行是否隐藏
            if ws.row_dimensions[row[0].row].hidden:
                continue
            dest_col = 1
            for cell in row:
                # 检查列是否隐藏
                if ws.column_dimensions[cell.column_letter].hidden:
                    continue
                new_ws.cell(dest_row, dest_col, cell.value)
                dest_col += 1
            dest_row += 1

    elif ext == ".xls":
        print("读取 XLS 格式文件（应用格式化）...")
        all_data = read_xls_with_formatting(src_path)
        
        # 写入新文件
        for row_idx, row_data in enumerate(all_data, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                new_ws.cell(row_idx, col_idx, cell_value)

    else:
        raise ValueError(f"不支持的文件格式: {ext}")

    new_wb.save(dest_path)
    print(f"\n✅ 可见数据已导出到 {dest_path}")


if __name__ == "__main__":
    # 你可以替换这里的路径进行测试
    src_file = "C:\\Users\\lupen\\Desktop\\aaa\\理化数据处理模板——飘红为统一要求，一张数据表中有且只有一列.xls"
    
    print("\n" + "="*80)
    print("测试: 读取.xls文件并应用格式化")
    print("="*80)
    
    # 读取并显示格式化后的数据
    all_data = read_xls_with_formatting(src_file)
    
    print("\n" + "="*80)
    print("前10行数据预览:")
    print("="*80)
    for i, row in enumerate(all_data[:10]):
        print(f"行{i}: {row[:10]}")  # 只显示前10列
    
    print("\n" + "="*80)
    print("导出到新的Excel文件...")
    print("="*80)
    copy_visible_excel(src_file, "formatted_output.xlsx")
